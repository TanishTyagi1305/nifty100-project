"""
peer_comparison_excel.py
--------------------------
Generates output/peer_comparison.xlsx -- 11 sheets, one per peer group,
with percentile-colored cells, benchmark row highlighted gold, and a
median summary row at the bottom of each sheet.
"""
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

METRICS = [
    "return_on_equity_pct", "net_profit_margin_pct", "debt_to_equity",
    "free_cash_flow_cr", "pat_cagr_5yr", "revenue_cagr_5yr",
    "eps_cagr_5yr", "interest_coverage", "asset_turnover",
]


def percentile_fill(pct):
    if pct is None:
        return None
    if pct >= 0.75:
        return GREEN
    if pct >= 0.25:
        return YELLOW
    return RED


def load_data():
    conn = sqlite3.connect("db/nifty100.db")
    percentiles = pd.read_sql("SELECT * FROM peer_percentiles", conn)
    companies = pd.read_sql("SELECT id AS company_id, company_name FROM companies", conn)
    peer_groups = pd.read_sql("SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn)
    conn.close()
    return percentiles, companies, peer_groups


def write_group_sheet(wb, group_name, group_companies, percentiles, companies, peer_groups):
    ws = wb.create_sheet(title=group_name[:31])

    header = ["company_id", "company_name"]
    for m in METRICS:
        header += [m, m + "_pctile"]
    ws.append(header)

    benchmark_row_num = None
    for company_id in group_companies:
        company_name_row = companies[companies["company_id"] == company_id]
        company_name = company_name_row["company_name"].iloc[0] if len(company_name_row) else company_id

        row = [company_id, company_name]
        for m in METRICS:
            match = percentiles[(percentiles["company_id"] == company_id) &
                                 (percentiles["peer_group_name"] == group_name) &
                                 (percentiles["metric"] == m)]
            if len(match):
                row += [match["value"].iloc[0], match["percentile_rank"].iloc[0]]
            else:
                row += [None, None]
        ws.append(row)

        is_bench = peer_groups[(peer_groups["company_id"] == company_id) &
                                (peer_groups["peer_group_name"] == group_name)]["is_benchmark"]
        if len(is_bench) and is_bench.iloc[0]:
            benchmark_row_num = ws.max_row

    # color the percentile columns
    for r in range(2, ws.max_row + 1):
        for i, m in enumerate(METRICS):
            pct_col = 3 + i * 2 + 1  # +1 because the pctile column follows the value column
            cell = ws.cell(row=r, column=pct_col)
            fill = percentile_fill(cell.value)
            if fill:
                cell.fill = fill

    # highlight benchmark row gold across all columns
    if benchmark_row_num:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=benchmark_row_num, column=c).fill = GOLD

    # summary median row at the bottom
    median_row = ["MEDIAN", ""]
    for m in METRICS:
        vals = percentiles[(percentiles["peer_group_name"] == group_name) &
                            (percentiles["metric"] == m)]["value"]
        median_row += [vals.median() if len(vals) else None, None]
    ws.append(median_row)


def main():
    percentiles, companies, peer_groups = load_data()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for group_name, group_df in peer_groups.groupby("peer_group_name"):
        write_group_sheet(wb, group_name, group_df["company_id"].tolist(),
                           percentiles, companies, peer_groups)

    wb.save("output/peer_comparison.xlsx")
    print(f"output/peer_comparison.xlsx written with {len(wb.sheetnames)} sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()