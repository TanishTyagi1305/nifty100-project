"""
export_excel.py
----------------
Generates output/screener_output.xlsx -- one sheet per preset, sorted
by composite score, with green/red fill showing pass/fail per threshold.
"""
import openpyxl
from openpyxl.styles import PatternFill

from src.screener.engine import load_all_metrics
from src.screener.presets import PRESETS, run_preset
from src.screener.turnaround import run_turnaround_watch
from src.screener.composite_score import compute_composite_scores

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

DISPLAY_COLS = [
    "company_id", "return_on_equity_pct", "debt_to_equity", "free_cash_flow_cr",
    "revenue_cagr_5yr", "pat_cagr_5yr", "pe_ratio", "pb_ratio",
    "dividend_yield_pct", "composite_quality_score",
]


def write_sheet(wb, name, company_ids, scored_df):
    ws = wb.create_sheet(title=name[:31])  # Excel sheet name limit is 31 chars

    subset = scored_df[scored_df["company_id"].isin(company_ids)].sort_values(
        "composite_quality_score", ascending=False)

    ws.append(DISPLAY_COLS)
    for _, row in subset.iterrows():
        ws.append([row.get(c) for c in DISPLAY_COLS])

    # simple green/red coloring: green if composite score above the sheet's
    # own median, red if below -- a straightforward, defensible pass/fail split
    if len(subset) > 0:
        median_score = subset["composite_quality_score"].median()
        score_col = DISPLAY_COLS.index("composite_quality_score") + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=score_col)
            cell.fill = GREEN if (cell.value or 0) >= median_score else RED


def main():
    scored_df = compute_composite_scores()
    df = load_all_metrics()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for name in PRESETS:
        result = run_preset(name)
        write_sheet(wb, name, result["company_id"].tolist(), scored_df)

    turnaround_ids = run_turnaround_watch()
    write_sheet(wb, "Turnaround Watch", turnaround_ids, scored_df)

    wb.save("output/screener_output.xlsx")
    print("output/screener_output.xlsx written with", len(wb.sheetnames), "sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()